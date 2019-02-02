from lxml import html
import openpyxl
import json
from time import gmtime, strftime

#writes data from ORDERS to trading log
def writeToExcelSheet(column_names, sheet, orders):
    item_names = column_names.split('\t')
    for i in range(len(item_names)):
        sheet.cell(row = 1, column = i + 1, value = item_names[i])
    for i in range(len(orders)):
        sheet.cell(row = i + 2, column = 1, value = orders[i].symbol)
        sheet.cell(row = i + 2, column = 2, value = orders[i].exitTime)
        sheet.cell(row = i + 2, column = 3, value = orders[i].exitPrice)
        sheet.cell(row = i + 2, column = 4, value = orders[i].size)
        sheet.cell(row = i + 2, column = 5, value = orders[i].direction)
        sheet.cell(row = i + 2, column = 6, value = orders[i].exitTime)
        sheet.cell(row = i + 2, column = 7, value = orders[i].exitPrice)
        sheet.cell(row = i + 2, column = 8, value = orders[i].strategy)

    sheet.cell(row = len(orders) + 5, column = 1, value = '')

#writes the content of testing dictionaries to respective JSON files
def writeTestResults(CALCULATED_EMAs, CANDLES, ORDERS, writeToExcel):
    with open('calculatedEMAs.json', 'w') as outfile:
        json.dump(CALCULATED_EMAs, outfile)

    with open('candles.json', 'w') as outfile:
        json.dump(CANDLES, outfile)

    #reset the orders file
    with open('orders.txt', 'w') as outfile:
        outfile.write('')

    with open('orders.txt', 'a') as outfile:
        outfile.write("Symbol" + '\t' + "Entry Time" + '\t' + "Entry Price" + '\t' + "Size" + '\t' + "Direction" + '\t' + "Exit Time" + '\t' + "Exit Price" + '\t' + "Strategy" + '\n')
        for i in range(len(ORDERS)):
            outfile.write(ORDERS[i].symbol + '\t' + ORDERS[i].entryTime + '\t' + str(ORDERS[i].entryPrice) + '\t' + str(ORDERS[i].size) + '\t'
                          + ORDERS[i].direction + '\t' + ORDERS[i].exitTime + '\t' + str(ORDERS[i].exitPrice) + '\t' + ORDERS[i].strategy + '\n')

    if (writeToExcel):
        time = strftime("%b %d, %Y", gmtime())
        wb = openpyxl.load_workbook(filename = 'trading log.xlsx')
        sheet = wb.create_sheet(time)
        column_names = 'Symbol	Entry Time	Entry Price	Size	Direction	Exit Time	Exit Price	Strategy'
        writeToExcelSheet(column_names, sheet, ORDERS)
        wb.save('trading log.xlsx')